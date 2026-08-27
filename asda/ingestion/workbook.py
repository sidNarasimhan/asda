"""Load CSV or Excel account lists the way a sales intern actually keeps them.

Continuation rows inherit the company. Extra sheets are skipped when they have
no person columns. One contact cell can become several people.
"""

from __future__ import annotations

import csv
import io
from pathlib import Path
from typing import Any, Iterable

from asda.ingestion.cleanup import clean_company, split_contacts, text

_ENCODINGS = ("utf-8-sig", "utf-8", "cp1252", "latin-1")
_HEADER_TOKENS = {
    "sl",
    "sno",
    "company",
    "account",
    "organization",
    "contact",
    "poc",
    "point",
    "person",
    "designation",
    "title",
    "linkedin",
    "phone",
    "mobile",
    "email",
    "mail",
    "remarks",
    "name",
    "website",
    "domain",
}
_PERSON_TOKENS = {
    "contact",
    "poc",
    "person",
    "email",
    "mail",
    "phone",
    "mobile",
    "linkedin",
    "designation",
    "title",
    "name",
    "client",
}
_COMPANY_KEYS = (
    "company_name",
    "company",
    "account_name",
    "organization",
    "account",
    "org",
)


def _clean_key(key: str) -> str:
    import re

    return re.sub(r"[^a-z0-9]+", "_", str(key or "").lower()).strip("_")


def _read_text(path: Path) -> str:
    data = path.read_bytes()
    for enc in _ENCODINGS:
        try:
            return data.decode(enc)
        except UnicodeDecodeError:
            continue
    return data.decode("utf-8", errors="replace")


def _delimiter(sample: str) -> str:
    header = sample.splitlines()[0] if sample else ""
    try:
        dialect = csv.Sniffer().sniff(sample[:8192], delimiters=",;\t|")
        return dialect.delimiter
    except csv.Error:
        pass
    counts = {",": header.count(","), ";": header.count(";"), "\t": header.count("\t"), "|": header.count("|")}
    best = max(counts, key=counts.get)
    return best if counts[best] else ","


def _matrix_csv(path: Path) -> list[list[str]]:
    sample = _read_text(path)
    if not sample.strip():
        return []
    delim = _delimiter(sample)
    return [list(r) for r in csv.reader(io.StringIO(sample), delimiter=delim)]


def _matrix_xlsx(path: Path) -> list[tuple[str, list[list[str]]]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError("Excel ingest needs openpyxl. pip install openpyxl") from exc
    wb = load_workbook(path, data_only=True, read_only=True)
    out: list[tuple[str, list[list[str]]]] = []
    for ws in wb.worksheets:
        rows: list[list[str]] = []
        for raw in ws.iter_rows(values_only=True):
            rows.append(["" if c is None else str(c) for c in raw])
        out.append((ws.title or "Sheet", rows))
    return out


def _header_index(matrix: list[list[str]]) -> int:
    best_i, best_score = 0, -1
    for i, row in enumerate(matrix[:25]):
        score = 0
        for cell in row:
            tokens = set(_clean_key(cell).split("_"))
            score += len(tokens & _HEADER_TOKENS)
        if score > best_score:
            best_i, best_score = i, score
    return best_i if best_score >= 2 else 0


def _has_person_columns(headers: list[str]) -> bool:
    keys = {_clean_key(h) for h in headers}
    tokens: set[str] = set()
    for k in keys:
        tokens.update(k.split("_"))
    return bool(tokens & _PERSON_TOKENS)


def _row_dict(headers: list[str], values: list[str]) -> dict[str, str]:
    row: dict[str, str] = {}
    extras: list[str] = []
    for i, header in enumerate(headers):
        name = text(header)
        if not name:
            val = text(values[i] if i < len(values) else "")
            if val:
                extras.append(val)
            continue
        val = text(values[i] if i < len(values) else "")
        if name in row and row[name]:
            continue
        row[name] = val
    if extras:
        row["_extra"] = " | ".join(extras)
    return row


def _company_from(row: dict[str, str]) -> str:
    for key, value in row.items():
        ck = _clean_key(key)
        if ck in _COMPANY_KEYS or ck.endswith("_name") and "company" in ck:
            cleaned = clean_company(value)
            if cleaned:
                return cleaned
    # first column sometimes *is* company when header is Account Name
    for key, value in row.items():
        ck = _clean_key(key)
        if ck in {"account_name", "company_name", "company"}:
            return clean_company(value)
    return ""


def _set_company(row: dict[str, str], company: str) -> None:
    for key in list(row):
        ck = _clean_key(key)
        if ck in _COMPANY_KEYS:
            row[key] = company
            return
    row["Company Name"] = company


def _has_person_values(row: dict[str, str]) -> bool:
    for key, value in row.items():
        if not text(value):
            continue
        ck = _clean_key(key)
        if ck in _COMPANY_KEYS or ck in {"sl", "sl_no", "sno", "remarks", "remark", "notes", "_extra"}:
            continue
        tokens = set(ck.split("_"))
        if tokens & _PERSON_TOKENS or ck in {"point_of_contact", "email_id"}:
            return True
    return False


def _forward_fill(rows: list[dict[str, str]]) -> list[dict[str, str]]:
    last = ""
    out: list[dict[str, str]] = []
    empty_run = 0
    for row in rows:
        company = _company_from(row)
        if company:
            last = company
        elif last:
            _set_company(row, last)
        if not _has_person_values(row):
            empty_run += 1
            if empty_run > 25:
                break
            continue
        empty_run = 0
        out.append(row)
    return out


def _expand(row: dict[str, str]) -> list[dict[str, str]]:
    contact_key = None
    for key in row:
        ck = _clean_key(key)
        if ck in {"contact_person_s", "contact_persons", "contact_person", "point_of_contact", "p_o_c", "poc"}:
            contact_key = key
            break
    if not contact_key:
        return [row]
    people = split_contacts(row[contact_key])
    if len(people) <= 1:
        if people:
            name, title = people[0]
            row[contact_key] = name
            if title and not any(_clean_key(k) in {"designation", "title", "job_title"} and text(v) for k, v in row.items()):
                row["Designation"] = title
        return [row]
    expanded: list[dict[str, str]] = []
    for name, title in people:
        clone = dict(row)
        clone[contact_key] = name
        if title:
            clone["Designation"] = title
        expanded.append(clone)
    return expanded


def tables_from(path: str | Path) -> list[tuple[str, list[dict[str, str]]]]:
    p = Path(path)
    suffix = p.suffix.lower()
    sheets: list[tuple[str, list[list[str]]]]
    if suffix in {".xlsx", ".xlsm", ".xltx", ".xls"}:
        sheets = _matrix_xlsx(p)
    else:
        sheets = [(p.stem, _matrix_csv(p))]
    out: list[tuple[str, list[dict[str, str]]]] = []
    for name, matrix in sheets:
        if not matrix:
            continue
        hi = _header_index(matrix)
        headers = [text(c) for c in matrix[hi]]
        if not any(headers) or not _has_person_columns(headers):
            continue
        rows = [_row_dict(headers, r) for r in matrix[hi + 1 :]]
        rows = _forward_fill(rows)
        expanded: list[dict[str, str]] = []
        for row in rows:
            expanded.extend(_expand(row))
        if expanded:
            out.append((name, expanded))
    return out


def iter_cleaned_rows(path: str | Path) -> Iterable[dict[str, str]]:
    for _sheet, rows in tables_from(path):
        for row in rows:
            yield row


def inspect(path: str | Path) -> dict[str, Any]:
    tables = tables_from(path)
    return {
        "path": str(path),
        "sheets": [{"name": n, "rows": len(rows)} for n, rows in tables],
        "rows": sum(len(r) for _, r in tables),
    }
