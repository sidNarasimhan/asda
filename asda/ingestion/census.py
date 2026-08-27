"""Count every line in uploaded account sheets vs unique people.

These workbooks repeat the same person across tabs. Consolidated Accounts
alone copies the list about seven times. Census makes that visible.
"""

from __future__ import annotations

import json
from pathlib import Path
from typing import Any

from asda.config import get_settings
from asda.ingestion.cleanup import (
    clean_company,
    company_core,
    extract_emails,
    extract_linkedin,
    extract_phones,
    parse_name_title,
    split_person_name,
    text,
    tidy_person_name,
)


def _report_path() -> Path:
    p = get_settings().data_dir / "ingest_report.json"
    p.parent.mkdir(parents=True, exist_ok=True)
    return p


def census_files(paths: list[str | Path]) -> dict[str, Any]:
    from openpyxl import load_workbook

    tabs: list[dict[str, Any]] = []
    union: dict[tuple[str, str], dict[str, Any]] = {}
    named_rows = 0
    nonempty = 0

    for path in paths:
        p = Path(path)
        if not p.exists():
            continue
        if p.suffix.lower() not in {".xlsx", ".xlsm", ".xls"}:
            continue
        wb = load_workbook(p, data_only=True, read_only=True)
        for ws in wb.worksheets:
            rows = list(ws.iter_rows(values_only=True))
            filled = sum(1 for r in rows if any(c not in (None, "") for c in r))
            nonempty += max(filled - 1, 0)
            people = _people(rows)
            named_rows += len(people)
            uniq = {(x["name"].lower(), company_core(x["company"]) or x["company"].lower()) for x in people}
            tabs.append(
                {
                    "file": p.name,
                    "sheet": ws.title,
                    "nonempty_rows": max(filled - 1, 0),
                    "named_people_rows": len(people),
                    "unique_people": len(uniq),
                }
            )
            for person in people:
                key = (person["name"].lower(), company_core(person["company"]) or person["company"].lower())
                prev = union.get(key)
                if not prev:
                    union[key] = person
                else:
                    if person["email"] and not prev["email"]:
                        prev["email"] = person["email"]
                    if person["linkedin"] and not prev["linkedin"]:
                        prev["linkedin"] = person["linkedin"]
                    if person["phone"] and not prev["phone"]:
                        prev["phone"] = person["phone"]
                    if len(person["company"]) > len(prev["company"]):
                        prev["company"] = person["company"]

    report = {
        "files": [str(p) for p in paths],
        "tabs": tabs,
        "nonempty_rows": nonempty,
        "named_people_rows": named_rows,
        "unique_people": len(union),
        "unique_with_email": sum(1 for p in union.values() if p["email"]),
        "unique_with_linkedin": sum(1 for p in union.values() if p["linkedin"]),
        "unique_with_phone": sum(1 for p in union.values() if p["phone"]),
        "note": (
            "Named rows are every line with a person. Unique people collapse copies. "
            "Consolidated Accounts repeats the same list several times."
        ),
    }
    _report_path().write_text(json.dumps(report, indent=2))
    return report


def load_report() -> dict[str, Any]:
    path = _report_path()
    if not path.exists():
        return {}
    try:
        return json.loads(path.read_text())
    except Exception:
        return {}


def _people(rows: list) -> list[dict[str, Any]]:
    if not rows:
        return []
    header = [text(h) for h in rows[0]]
    idx = {h.lower(): i for i, h in enumerate(header) if h}

    def col(*names: str) -> int | None:
        for n in names:
            if n.lower() in idx:
                return idx[n.lower()]
        return None

    co_i = col("Company Name", "Account Name", "Company", "Organization", "COMPANY")
    poc_i = col(
        "Point of Contact",
        "Contact Person(s)",
        "Contact Person",
        "Contact Name",
        "Contact",
        "Name",
        "P.O.C",
        "Client Name",
        "NAME",
    )
    em_i = col("Email ID", "Email", "Work Email", "Personal Email", "E MAIL ID", "Mail ID")
    li_i = col("LinkedIn", "LinkedIn Profile")
    ph_i = col("Phone Number", "Phone", "Contact No", "Phone 1", "CONTACT NO", "Number")
    last = ""
    out: list[dict[str, Any]] = []
    for row in rows[1:]:
        if not any(c not in (None, "") for c in row):
            continue

        def cell(i: int | None) -> str:
            if i is None or i >= len(row):
                return ""
            return text(row[i])

        co = clean_company(cell(co_i)) or last
        if clean_company(cell(co_i)):
            last = clean_company(cell(co_i))
        poc = cell(poc_i)
        if not poc:
            continue
        name, _title = parse_name_title(poc)
        first, lastn = tidy_person_name(*split_person_name(name))
        full = f"{first} {lastn}".strip()
        if not full:
            continue
        emails = extract_emails(cell(em_i))
        out.append(
            {
                "name": full,
                "company": co,
                "email": emails[0] if emails else "",
                "linkedin": extract_linkedin(cell(li_i) or poc),
                "phone": extract_phones(cell(ph_i)),
            }
        )
    return out
